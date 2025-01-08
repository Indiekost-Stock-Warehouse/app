import customtkinter as ctk
from tkinter import messagebox
import os
from openpyxl import Workbook, load_workbook
from PIL import Image
import sqlite3
from datetime import datetime
import user.stok as userstok
import user.transaksi as usertransaksi

# Koneksi ke database SQLite
def get_item_names():
    conn = sqlite3.connect("db_p3l.db")
    cursor = conn.cursor()
    cursor.execute("SELECT nama FROM gudang")
    items = [row[0] for row in cursor.fetchall()]
    conn.close()
    return items

# Fungsi untuk menyimpan data ke file Excel
def save_to_excel(directory, filename, data):
    # Pastikan direktori tujuan ada
    os.makedirs(directory, exist_ok=True)

    # Path lengkap file Excel
    filepath = os.path.join(directory, filename)

    # Periksa apakah file sudah ada
    if os.path.exists(filepath):
        # Jika ada, buka file Excel
        workbook = load_workbook(filepath)
        sheet = workbook.active
    else:
        # Jika belum ada, buat file baru
        workbook = Workbook()
        sheet = workbook.active
        # Tambahkan header
        sheet.append(["No", "Jam Pemesanan", "Barang Pesanan", "Jumlah"])

    # Tambahkan data baru ke file
    no = sheet.max_row  # Menentukan nomor baris terakhir
    sheet.append([no, data["jam_pemesanan"], data["barang_pesanan"], data["jumlah"]])

    # Simpan file Excel
    workbook.save(filepath)

# Images
logo_img_data = Image.open("./Asset/Image/logo.png")
logo_img = ctk.CTkImage(dark_image=logo_img_data, light_image=logo_img_data, size=(77.68, 85.42))

analytics_img_data = Image.open("./Asset/Image/analytics_icon.png")
analytics_img = ctk.CTkImage(dark_image=analytics_img_data, light_image=analytics_img_data)

package_img_data = Image.open("./Asset/Image/package_icon.png")
package_img = ctk.CTkImage(dark_image=package_img_data, light_image=package_img_data)

list_img_data = Image.open("./Asset/Image/list_icon.png")
list_img = ctk.CTkImage(dark_image=list_img_data, light_image=list_img_data)

returns_img_data = Image.open("./Asset/Image/returns_icon.png")
returns_img = ctk.CTkImage(dark_image=returns_img_data, light_image=returns_img_data)

settings_img_data = Image.open("./Asset/Image/settings_icon.png")
settings_img = ctk.CTkImage(dark_image=settings_img_data, light_image=settings_img_data)

person_img_data = Image.open("./Asset/Image/person_icon.png")
person_img = ctk.CTkImage(dark_image=person_img_data, light_image=person_img_data)

def lihat_order(app):
    # App Windows GUI
    for widget in app.winfo_children():
        widget.destroy()

    app.title("Indiekost Order")

    sidebar_frame = ctk.CTkFrame(master=app, fg_color="#2A8C55",  width=176, height=650, corner_radius=0)
    sidebar_frame.pack_propagate(0)
    sidebar_frame.pack(fill="y", anchor="w", side="left")

    ctk.CTkLabel(master=sidebar_frame, text="", image=logo_img).pack(pady=(38, 0), anchor="center")

    ctk.CTkButton(master=sidebar_frame, image=package_img, text="Pesanan", fg_color="#fff", font=("Arial Bold", 14), text_color="#2A8C55", hover_color="#eee", anchor="w").pack(anchor="center", ipady=5, pady=(16, 0))

    ctk.CTkButton(master=sidebar_frame, image=list_img, text="Stock", command=lambda: userstok.main_ui(app), fg_color="transparent", font=("Arial Bold", 14), hover_color="#207244", anchor="w").pack(anchor="center", ipady=5, pady=(16, 0))

    ctk.CTkButton(master=sidebar_frame, image=returns_img, text="Transaksi", command=lambda: usertransaksi.main(app), fg_color="transparent", font=("Arial Bold", 14), hover_color="#207244", anchor="w").pack(anchor="center", ipady=5, pady=(16, 0))

    main_view = ctk.CTkFrame(master=app, fg_color="#fff",  width=680, height=650, corner_radius=0)
    main_view.pack_propagate(0)
    main_view.pack(side="left")

    ctk.CTkLabel(master=main_view, text="Order", font=("Arial Black", 25), text_color="#2A8C55").pack(anchor="nw", pady=(29,0), padx=27)
    ctk.CTkLabel(master=main_view, text="Barang Pesanan", font=("Arial Bold", 17), text_color="#52A476").pack(anchor="nw", pady=(25,0), padx=27)

    # Dropdown untuk memilih nama item dari database
    item_names = get_item_names()
    selected_item = ctk.StringVar(value="Pilih Item")
    ctk.CTkOptionMenu(master=main_view, variable=selected_item, values=item_names, fg_color="#F0F0F0", text_color="#52A476").pack(fill="x", pady=(12,0), padx=27, ipady=10)

    grid = ctk.CTkFrame(master=main_view, fg_color="transparent")
    grid.pack(fill="both", padx=27, pady=(31,0))

    # Fungsi untuk memperbarui stok di database
    def update_stock(item_name, jumlah_pesanan, operation):
        conn = sqlite3.connect("db_p3l.db")
        cursor = conn.cursor()

        # Ambil stok saat ini
        cursor.execute("SELECT stok FROM gudang WHERE nama = ?", (item_name,))
        result = cursor.fetchone()

        if result:
            current_stock = result[0]
            if operation == "add":  # Menambah stok (pembelian)
                new_stock = current_stock + int(jumlah_pesanan)
            elif operation == "subtract":  # Mengurangi stok (penjualan)
                if current_stock >= int(jumlah_pesanan):  # Pastikan stok cukup
                    new_stock = current_stock - int(jumlah_pesanan)
                else:
                    print("Stok tidak cukup!")
                    conn.close()
                    return False  # Gagal karena stok tidak mencukupi
            
            # Perbarui stok di database
            cursor.execute("UPDATE gudang SET stok = ? WHERE nama = ?", (new_stock, item_name))
            conn.commit()
            conn.close()
            return True  # Berhasil
        else:
            print("Item tidak ditemukan di database!")
            conn.close()
            return False  # Gagal karena item tidak ditemukan

    # Fungsi untuk pembelian
    def beli():
        item_name = selected_item.get()
        jumlah_pesanan = jumlah_pesanan_entry.get()

        if jumlah_pesanan.isdigit() and int(jumlah_pesanan) > 0:
            success = update_stock(item_name, jumlah_pesanan, "add")
            if success:
                messagebox.showinfo("Buy Item", f"Berhasil Membeli {item_name} untuk stok sebanyak {jumlah_pesanan}.")
                # Simpan data ke spreadsheet
                save_to_excel(
                    directory="./report/buy",
                    filename=datetime.now().strftime("%Y-%m-%d") + ".xlsx",
                    data={
                        "jam_pemesanan": datetime.now().strftime("%H:%M:%S"),
                        "barang_pesanan": item_name,
                        "jumlah": jumlah_pesanan,
                    },
                )
            else:
                messagebox.showerror("Error", f"Gagal menambah stok untuk {item_name}.")
        else:
            messagebox.showerror("Error", "Jumlah pesanan harus berupa angka positif!")

    # Fungsi untuk penjualan
    def jual():
        item_name = selected_item.get()
        jumlah_pesanan = jumlah_pesanan_entry.get()

        if jumlah_pesanan.isdigit() and int(jumlah_pesanan) > 0:
            success = update_stock(item_name, jumlah_pesanan, "subtract")
            if success:
                messagebox.showinfo("Sell Item", f"Behasil Menjual {item_name} sebanyak {jumlah_pesanan}.")
                # Simpan data ke spreadsheet
                save_to_excel(
                    directory="./report/sell",
                    filename=datetime.now().strftime("%Y-%m-%d") + ".xlsx",
                    data={
                        "jam_pemesanan": datetime.now().strftime("%H:%M:%S"),
                        "barang_pesanan": item_name,
                        "jumlah": jumlah_pesanan,
                    },
                )
            else:
                messagebox.showerror("Error", f"Gagal Menjual {item_name}.")
        else:
            messagebox.showerror("Error", "Jumlah pesanan harus berupa angka positif!")


    ctk.CTkLabel(master=grid, text="Tanggal Pemesanan", font=("Arial Bold", 17), text_color="#52A476", justify="left").grid(row=0, column=0, sticky="w")
    
    # Entry otomatis untuk tanggal dan waktu saat ini
    current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    tgl_pemesanan_entry = ctk.CTkEntry(master=grid, fg_color="#F0F0F0", border_width=0, width=300)
    tgl_pemesanan_entry.insert(0, current_datetime)
    tgl_pemesanan_entry.configure(state="disabled")  # Nonaktifkan agar tidak bisa diubah
    tgl_pemesanan_entry.grid(row=1, column=0, ipady=10)

    ctk.CTkLabel(master=grid, text="Jumlah Pesanan", font=("Arial Bold", 17), text_color="#52A476", justify="left").grid(row=0, column=1, sticky="w", padx=(25,0))
    jumlah_pesanan_entry = ctk.CTkEntry(master=grid, fg_color="#F0F0F0", border_width=0, width=300)
    jumlah_pesanan_entry.grid(row=1, column=1, ipady=10, padx=(24,0))

    actions = ctk.CTkFrame(master=main_view, fg_color="transparent")
    actions.pack(fill="both")

    # Tombol "Beli" (Pembelian)
    ctk.CTkButton(
        master=actions, 
        text="Beli", 
        width=300, 
        fg_color="transparent", 
        font=("Arial Bold", 17), 
        border_color="#2A8C55", 
        hover_color="#eee", 
        border_width=2, 
        text_color="#2A8C55", 
        command=beli
    ).pack(side="left", anchor="sw", pady=(30, 0), padx=(27, 24))

    # Tombol "Jual" (Penjualan)
    ctk.CTkButton(
        master=actions, 
        text="Jual", 
        width=300, 
        font=("Arial Bold", 17), 
        hover_color="#207244", 
        fg_color="#2A8C55", 
        text_color="#fff", 
        command=jual
    ).pack(side="left", anchor="se", pady=(30, 0), padx=(0, 27))

    app.mainloop()